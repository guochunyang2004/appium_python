# -*- coding:utf-8 -*-
import time

from appium import webdriver

# from pymongo import MongoClient
import threading

import openpyxl
from openpyxl.drawing.image import Image
import os
import re #正则
basePath = os.path.abspath('.')#获得当前工作目录
imgPath=basePath+u'/res/img' 
sheetImgName = 'SheetImg'

maxPage = 10  # 最大滑动次数
# --------------------打开excel文件,获取工作簿对象----------------------------
filePath = basePath+'/res/【金领冠悠滋小羊】小红书资源list.xlsx'
wb = openpyxl.load_workbook(filePath)
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单
print('----Excel {}行 {}列-----'.format(ws.max_row, ws.max_column))  
# android手机配置 
caps = {
    "platformName": "Android",
    "platformVersion": "9",  #----------安卓版本号，设置>系统>关于手机>Android 版本-----------
    "deviceName": "5JPDU17826008316", #-------------设备名，查看命令：adb devices-------------
    "appPackage": "com.xingin.xhs",    
    "appActivity": ".activity.SplashActivity",
    "noReset": True, # 免登陆TRUE
    "codeKeyboard": True # 解决不能输入中文的问题
}
driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', caps)


def getSize():
    x = driver.get_window_size()['width']
    y = driver.get_window_size()['height']
    return (x, y)

def swipeUp():
    l = getSize()
    x1 = int(l[0] * 0.5) # x坐标
    y1 = int(l[1] * 0.75) # 起始y坐标
    y2 = int(l[1] * 0.25) # 终点y坐标
    driver.swipe(x1, y1, x1, y2)

def swipeDown():
    l = getSize()
    x1 = int(l[0] * 0.5) # x坐标
    y1 = int(l[1] * 0.25) # 起始y坐标
    y2 = int(l[1] * 0.75) # 终点y坐标
    driver.swipe(x1, y1, x1, y2)
    
def swipLeft(t=500, n=1):
    '''向左滑动屏幕'''
    l = driver.get_window_size()
    x1 = l['width'] * 0.75
    y1 = l['height'] * 0.5
    x2 = l['width'] * 0.25
    for i in range(n):
        driver.swipe(x1, y1, x2, y1, t)
        
def swipRight(t=500, n=1):
    '''向右滑动屏幕'''
    l = driver.get_window_size()
    x1 = l['width'] * 0.25
    y1 = l['height'] * 0.5
    x2 = l['width'] * 0.75
    for i in range(n):
        driver.swipe(x1, y1, x2, y1, t)

def data():
    title = driver.find_element_by_id("com.xingin.xhs:id/bdb").text
    content = driver.find_element_by_id("com.xingin.xhs:id/bbo").text
    print("标题——>", title)
    print("内容——>", content)
    swipeUp()
    # swipeUp(3500)


    try:
        collect = driver.find_element_by_id("com.xingin.xhs:id/bcc").text
    except Exception as e:
        print(e)
        collect = '无收藏'
    try:
        like_num = driver.find_element_by_id("com.xingin.xhs:id/bd3").text
    except Exception as e:
        print(e)
        like_num = '无点赞'
    try:
        comments = driver.find_elements_by_id("com.xingin.xhs:id/bbo")[1].text
    except Exception as e:
        print(e)
        comments = '无评论'
    data = {'title': title, 'content': content, 'comments': comments, 'collect': collect, 'like_num': like_num}
    print('存入数据库前。。。。。。。。。。。。')
    # cur.insert(data)
    print(data)
    print('存入成功。。。。。。。。。。')

def alertClick():
    alertTitle = driver.find_elements_by_id('android:id/alertTitle')
    if len(alertTitle)>0:
        if alertTitle[0].text.strip()=="新版发布":
            # 取消
            cancelButton = driver.find_elements_by_xpath("/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.Button[1]")
            if len(lencancelButton)>0:
                cancelButton[0].click()
                time.sleep(1)
# 返回
def clickBack():
    backButton = driver.find_elements_by_id("com.xingin.xhs:id/k1")
    if len(backButton)>0:
        backButton[0].click()
        return True
    return False

# 复制链接
def clickCopyShareLink():
    # com.xingin.xhs:id/c_w /hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.LinearLayout/androidx.recyclerview.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout
    # com.xingin.xhs:id/anw /hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.LinearLayout/androidx.recyclerview.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.ImageView
    # com.xingin.xhs:id/cqq  
    shareLink = driver.find_elements_by_xpath('/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.LinearLayout/androidx.recyclerview.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.ImageView')
    if len(shareLink)>0:
        shareLink[0].click()
        time.sleep(1)
        #print("----shareLink.click()----")
        str = driver.get_clipboard_text() #获取粘贴板内容
        return str
    return ''

def findUrl(string): 
    # findall() 查找匹配正则表达式的字符串
    #url = re.findall('https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+', string)
    url = re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+',string)
    if len(url)>0:
        return url[0]
    return ''

# 截图
def screenshot(r):    
    driver.get_screenshot_as_file(u'{}/img_{}.png'.format(imgPath,r))

# 创建新sheet，已存在删除
def createImgSheetAndGet():   
    sheets = wb.sheetnames
    if sheetImgName in sheets:
        # 删除sheet
        sheetImg = wb.get_sheet_by_name(sheetImgName)
        wb.remove(sheetImg)    
    sheetImg = wb.create_sheet(sheetImgName)
    return sheetImg
    
def addImg(sheetImg,r):
    path = u'{}/img_{}.png'.format(imgPath,r)
    if not os.path.exists(path):
        return False
    #ws.cell(row=r, column=1).value=r
    # 设置文字图片单元格的行高列宽
    column_width=255
    row_height=450   # 设置行高，该设置的行高与excel文件中设置的行高值是一样的
    # r = 2 #3行
    colName = u'B'
    # 下面代码中的[]括号中可以输入colName或者colName
    sheetImg.row_dimensions[r].height=row_height*0.8   # 修改行3的行高
    sheetImg.column_dimensions[colName].width=22   # 设置字符宽度
    img = Image(path)
    
    newSize=(column_width,row_height)
    img.width,img.height=newSize    # 这两个属性分别是对应添加图片的宽高
    sheetImg.add_image(img,colName+str(r)) # 向M列中的单元格内指定添加图片
    # -------写入图片 结束---------
    return True

def delImg(r):
    path = u'{}/img_{}.png'.format(imgPath,r)
    if os.path.exists(path):
        os.remove(path)

def getRedBook():
    # 链接app
    #driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', caps)
    time.sleep(2)
    sheetImg = createImgSheetAndGet() # 重置图片sheet
    rowIndex = 0
    # keys = [u"孩子营养好吸收 推荐喝羊奶粉悠滋小羊测评",u'悠滋小羊测评|带你认识真正好的羊奶粉！']
    #keys = [u'悠滋小羊测评|带你认识真正好的羊奶粉！',u"孩子营养好吸收 推荐喝羊奶粉悠滋小羊测评",u"越来越多妈妈选羊奶粉，真相原来是这样",u"断奶期妈妈选择｜推荐这款羊奶粉"]
    #for searchKey in keys:
    # 遍历Excel
    for col in ws.iter_rows(min_row=3, max_row=ws.max_row,min_col=11, max_col=12):
        rowIndex += 1
        if col[0].value and col[0].value!='删除':   
            searchKey = col[0].value
        else:
            continue
        alertClick()
        time.sleep(1)
        swipeDown()
        time.sleep(1)
        # 首页假输入框
        #t = driver.find_element_by_id("com.xingin.xhs:id/we").click()
        # 查找发现页搜索框
        searchView = driver.find_elements_by_xpath("/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/androidx.drawerlayout.widget.DrawerLayout/android.widget.LinearLayout/android.widget.RelativeLayout/androidx.viewpager.widget.ViewPager/android.widget.LinearLayout/android.widget.FrameLayout[1]/android.widget.FrameLayout/android.widget.FrameLayout[2]/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout")
        if len(searchView)==0:
            #查找页搜索框
            searchView = driver.find_elements_by_id("com.xingin.xhs:id/bat")
            if len(searchView)==0:
                print("未找到搜索框")
                break
        searchView[0].click()
        time.sleep(2)
        # 真实输入框
        text = driver.find_element_by_id("com.xingin.xhs:id/bab")
        time.sleep(1)
        #searchKey = u"孩子营养好吸收 推荐喝羊奶粉悠滋小羊测评"
        text.send_keys(searchKey)
        time.sleep(1)
        # 搜索按键
        driver.find_element_by_id("com.xingin.xhs:id/bae").click()
        time.sleep(2)
        # 向下滑动，刷新数据
        swipeDown()
        # swipeDown(500)
        time.sleep(2)
        cnt = maxPage
        isFound = False
        print("begin======-------=======")
        while cnt>0:        
            try:
                titleList = driver.find_elements_by_xpath('//*[@resource-id="com.xingin.xhs:id/b96"]')
                # title = driver.find_element_by_xpath("/android.widget.TextView[@resource-id='com.xingin.xhs:id/b96']")
                #                                    "/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.viewpager.widget.ViewPager/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[2]/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.TextView"
                # title = driver.find_element_by_xpath("/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.viewpager.widget.ViewPager/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout[1]/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.TextView")
                #if titleList:
                    #print("titleList",titleList.text)
                #titleList = driver.find_elements_by_xpath("/hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/androidx.viewpager.widget.ViewPager/android.widget.FrameLayout/androidx.recyclerview.widget.RecyclerView/android.widget.FrameLayout")
                #print("len(titleList)：",len(titleList))
                for item in titleList:
                    #title = item.find_element_by_xpath("./android.widget.LinearLayout/android.widget.LinearLayout/android.widget.TextView")
                    #if item.text and item.text.strip()==searchKey.strip() :
                    #print(item.text)
                    if item.text and item.text.find(searchKey.strip())!=-1 :                        
                        isFound = True                
                        screenshot(rowIndex) # 截图
                        time.sleep(1)
                        addImg(sheetImg,rowIndex) # 添加图片                        
                        wb.save(filePath)
                        ##delImg(rowIndex) # 删除图片
                        
                        item.click()
                        time.sleep(2)
                        # /hierarchy/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.RelativeLayout/android.widget.ImageView[2]
                        # 点击分享按钮
                        driver.find_element_by_id('com.xingin.xhs:id/moreOperateIV').click()
                        time.sleep(1)
                        link = clickCopyShareLink()#复制链接
                        link = findUrl(link) # 正则替换链接 
                        col[1].value = link    
                        wb.save(filePath)
                        
                        clickBack()  # 返回上一页  
                        print("---searchKey----",cnt,searchKey,"len:",len(titleList),link)            
                        break
                # for titleList end                
           
                if isFound:
                    break
                swipeUp() # 上滑
                time.sleep(1)
                cnt=cnt-1 
                
            except Exception as r:
                print('错误 %s' %r)
                time.sleep(2)
                swipeUp()
                # swipeUp(350)
        # while end
    #for keys end
    
if __name__ == "__main__":
    getRedBook()    