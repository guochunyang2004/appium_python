# -*- coding:utf-8 -*-
import openpyxl
import os
from openpyxl.drawing.image import Image
basePath = os.path.abspath('.')#获得当前工作目录
imgPath=basePath+u'/res/img'

# 打开excel文件,获取工作簿对象
filePath = basePath+'/res/【金领冠悠滋小羊】小红书资源list.xlsx'

wb = openpyxl.load_workbook(filePath)
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单


print('{}行 {}列'.format(ws.max_row, ws.max_column))   
rowIndex = 1
# 打印3-max_row行，11列中的内容:
for col in ws.iter_rows(min_row=3, max_row=ws.max_row,min_col=11, max_col=12): 
    if col[0].value and col[0].value!='删除':
        col[1].value = 'http://'
        print('{} {}'.format(rowIndex,col[0].value))
    else:
        col[1].value = '无'
    rowIndex += 1

# -------写入图片---------

# 创建新sheet
sheetImgName = 'SheetImg'
sheets = wb.sheetnames
if sheetImgName in sheets:
    # 删除sheet
    sheetImg = wb.get_sheet_by_name(sheetImgName)
    wb.remove(sheetImg)    
sheetImg = wb.create_sheet(sheetImgName)

def addImg(r):
    #ws.cell(row=r, column=1).value=r
    # 设置文字图片单元格的行高列宽
    column_width=170
    row_height=300   # 设置行高，该设置的行高与excel文件中设置的行高值是一样的
    # r = 2 #3行
    colName = u'B'
    # 下面代码中的[]括号中可以输入colName或者colName
    sheetImg.row_dimensions[r].height=row_height*0.8   # 修改行3的行高
    sheetImg.column_dimensions[colName].width=22   # 设置字符宽度
    img = Image(imgPath+u'/test.png')
    newSize=(column_width,row_height)
    img.width,img.height=newSize    # 这两个属性分别是对应添加图片的宽高
    sheetImg.add_image(img,colName+str(r)) # 向M列中的单元格内指定添加图片
    
    # -------写入图片 结束---------
def delImg(r):
    path = u'{}/test1.png'.format(imgPath)
    if os.path.exists(path):
        os.remove(path)
        
addImg(2)
addImg(3)

delImg(1)

wb.save(filePath)

# 从工作薄中获取一个表单(sheet)对象
#sheets = wb.sheetnames
#print(sheets, type(sheets))
 
# 创建一个表单
# wb.get_sheet
# mySheet = wb.create_sheet('mySheet')
# print(wb.sheetnames)
 
# 获取指定的表单
# sheet1 = wb.get_sheet_by_name('Sheet1')
# print(sheet1.title)
#sheet4 = wb['mySheet']
#for sheet in wb:
#   print(sheet.title)

# print(ws)
# print(ws['A1']) # 获取A列的第一个对象
# print(ws['A1'].value)


# print(ws.cell(row=1, column=2)) # 获取第一行第二列的单元格
# print(ws.cell(row=1, column=2).value)
# for i in range(1, 8, 2): #  获取1,3,4,7 行第二列的值
#     print(i, ws.cell(row=i, column=2).value)

# 遍历行和列中单元格的值
# rowIndex = 0
# for row in ws.iter_rows(min_row=1, max_row=33, max_col=11):
#     str = ''
#     rowIndex += 1
#     for cell in row:
#         str += '{} {},'.format(rowIndex,cell.value)
#     print(str)

    
# colC = ws['C'] # 获取整个C列
# print(colC)
# row6 = ws[6]   # 获取第6行
# print(row6,type(row6))

# col_range = ws['B:C']
# row_range = ws[2:6] 
# 遍历行和列中单元格的值　　
# for col in col_range: # 打印BC两列单元格中的值内容
#     for cell in col:
#         print(cell.value)
 
# for row in row_range: # 打印 2-5行中所有单元格中的值
#     for cell in row:
#         print(cell.value)
 
# for row in ws.iter_rows(min_row=1, max_row=2, max_col=2): # 打印1-2行，1-2列中的内容
#     for cell in row:
#         print(cell.value)
     


# # 把数字转换成字母
# print(get_column_letter(2), get_column_letter(47), get_column_letter(900))
 
# # 把字母转换成数字
# print(column_index_from_string('AAH'))